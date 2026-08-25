from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
section_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
body_font = Font(name='Arial', size=10)
note_font = Font(name='Arial', size=9, italic=True, color='666666')

header_fill = PatternFill('solid', fgColor='2D2D2D')
spider_fill = PatternFill('solid', fgColor='5B21B6')
collector_fill = PatternFill('solid', fgColor='0D9488')
row_even = PatternFill('solid', fgColor='F5F5F5')
row_odd = PatternFill('solid', fgColor='FFFFFF')

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap = Alignment(vertical='top', wrap_text=True)

cols = ['Test #', 'Extension', 'Category', 'Test Case', 'Steps', 'Expected Result', 'Status', 'Notes']
col_widths = [8, 12, 16, 22, 40, 35, 12, 30]

browsers = {
    'Chrome': 'Standard MV3. Test via chrome://extensions (Developer mode). Primary target platform.',
    'Firefox': 'Uses browser.* API namespace. Test via about:debugging. May need gecko ID in manifest. Check MV2/MV3 compatibility.',
    'Edge': 'Same MV3 as Chrome. Publish via Edge Add-ons store (or sideload). Test via edge://extensions.'
}

spider_tests = [
    ['S-01', 'Spider', 'Launch', 'Basic Launch',
     'Open ChatGPT in browser. Click Spider icon in toolbar.',
     'Popup loads without errors. Shows dashboard or ready state.'],
    ['S-02', 'Spider', 'Export', 'Single Conversation Export',
     'Navigate to one ChatGPT conversation. Use single-export button.',
     'JSON file downloads. Contains messages array with correct roles and content.'],
    ['S-03', 'Spider', 'Batch', 'Batch Scrape (Small)',
     'Start batch scrape targeting 5-10 conversations.',
     'Progress counter updates in popup. Each conversation scraped sequentially.'],
    ['S-04', 'Spider', 'Batch', 'Pause / Resume',
     'Start batch scrape. Click Pause mid-scrape. Close popup. Reopen popup. Click Resume.',
     'Paused state persists across popup close/reopen. Resume continues from where it stopped.'],
    ['S-05', 'Spider', 'Batch', 'Partial Save',
     'During batch scrape, trigger partial save / download.',
     'JSON downloads with all conversations scraped so far. Scrape continues after save.'],
    ['S-06', 'Spider', 'Recovery', 'Extension Reload Recovery',
     'Start scrape. Go to extensions page and reload Spider. Reopen popup.',
     'Session state recovered or partial data not lost. Known bug area - document behavior.'],
    ['S-07', 'Spider', 'Edge Case', 'Empty Conversation',
     'Navigate to a conversation with no assistant replies. Attempt scrape.',
     'Does not crash. Exports empty or minimal JSON gracefully.'],
    ['S-08', 'Spider', 'Edge Case', 'Long Conversation',
     'Scrape a conversation with 50+ messages.',
     'All messages captured. No truncation. Correct ordering maintained.'],
    ['S-09', 'Spider', 'Validation', 'Output JSON Validation',
     'Open exported JSON in editor. Spot-check 5+ conversations.',
     'Valid JSON. Correct message ordering, roles (user/assistant/system). No HTML artifacts.'],
    ['S-10', 'Spider', 'Error', 'Navigate Away Mid-Scrape',
     'Start batch scrape. Navigate active tab away from ChatGPT.',
     'Graceful error or pause. No crash. Can recover when returning to ChatGPT.'],
    ['S-11', 'Spider', 'Concurrency', 'Multiple Tabs',
     'Open ChatGPT in 2+ tabs. Start scrape from one.',
     'No conflicts between tabs. Scrape uses correct tab.'],
    ['S-12', 'Spider', 'Dev', 'Console Clean',
     'Run through all tests with DevTools console open.',
     'No errors or unexpected warnings in console.'],
]

collector_tests = [
    ['C-01', 'Collector', 'Launch', 'Basic Launch',
     'Open ChatGPT conversation containing images. Click Collector icon.',
     'Popup loads. Shows image count or scan status.'],
    ['C-02', 'Collector', 'Detection', 'Image Detection (DALL-E)',
     'Open conversation with DALL-E generated images.',
     'All DALL-E images found and listed with thumbnails.'],
    ['C-03', 'Collector', 'Download', 'Single Image Download',
     'Click download on one detected image.',
     'Image saves correctly. Correct format (PNG/JPG). Not corrupted. Opens normally.'],
    ['C-04', 'Collector', 'Download', 'Batch Download All',
     'Click "Download All" with multiple images detected.',
     'All images save with unique filenames. No overwrites. Correct formats.'],
    ['C-05', 'Collector', 'Detection', 'Image Types Variety',
     'Test conversations with: DALL-E, user uploads, charts/diagrams, code screenshots.',
     'All image types detected and downloadable.'],
    ['C-06', 'Collector', 'Edge Case', 'No Images',
     'Open a text-only conversation. Click Collector.',
     'Shows "no images found" or similar message. No crash.'],
    ['C-07', 'Collector', 'Scale', 'Large Conversation (20+ images)',
     'Open conversation with 20+ images.',
     'All images found. Scroll/pagination works if needed. No performance issues.'],
    ['C-08', 'Collector', 'Edge Case', 'Duplicate Images',
     'Conversation where same image appears multiple times.',
     'Behavior documented: deduplicates or downloads each instance.'],
    ['C-09', 'Collector', 'Error', 'Expired Image URLs',
     'Open a very old conversation where image URLs have expired.',
     'Graceful error message per image. No crash. Working images still downloadable.'],
    ['C-10', 'Collector', 'Download', 'Download Location',
     'Download images and check save location.',
     'Images save to default Downloads folder or browser-configured location.'],
    ['C-11', 'Collector', 'Dev', 'Console Clean',
     'Run through all tests with DevTools console open.',
     'No errors or unexpected warnings in console.'],
]

for idx, (browser, browser_note) in enumerate(browsers.items()):
    if idx == 0:
        ws = wb.active
        ws.title = browser
    else:
        ws = wb.create_sheet(browser)

    # Browser note row
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{browser} Testing Notes: {browser_note}'
    ws['A1'].font = Font(name='Arial', size=10, italic=True, color='444444')
    ws['A1'].fill = PatternFill('solid', fgColor='FFF3CD')
    ws['A1'].alignment = wrap

    # Header row
    for c, (col_name, width) in enumerate(zip(cols, col_widths), 1):
        cell = ws.cell(row=2, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c)].width = width

    row = 3

    # Spider section header
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = 'ANTHILL SPIDER -- Conversation Exporter'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = spider_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    for i, test in enumerate(spider_tests):
        fill = row_even if i % 2 == 0 else row_odd
        for c, val in enumerate(test, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = wrap
            cell.border = thin_border
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=8).font = note_font
        ws.cell(row=row, column=8).fill = fill
        ws.cell(row=row, column=8).border = thin_border
        row += 1

    row += 1

    # Collector section header
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = 'ANTHILL COLLECTOR -- Image Extractor'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = collector_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    for i, test in enumerate(collector_tests):
        fill = row_even if i % 2 == 0 else row_odd
        for c, val in enumerate(test, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = wrap
            cell.border = thin_border
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=8).font = note_font
        ws.cell(row=row, column=8).fill = fill
        ws.cell(row=row, column=8).border = thin_border
        row += 1

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25
    for r in range(3, row + 1):
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:H{row - 1}'

wb.save(r'D:\repos\anthill\Anthill_Test_Plan.xlsx')
print('Saved D:\\repos\\anthill\\Anthill_Test_Plan.xlsx')
